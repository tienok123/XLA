"""
Tiện ích xử lý file - Bỏ JSON, chỉ lưu Excel và ảnh
"""
import os
import cv2
import numpy as np
from datetime import datetime
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox

from openpyxl.utils import get_column_letter

from core.config import OUTPUTS_DIR, PRODUCT_NAMES_VI, QUALITY_NAMES_VI


def save_results(processed_image, detections, settings, original_image_path=None):
    """
    Lưu kết quả phân tích ảnh: ảnh processed, Excel, overlay
    ĐÃ BỎ JSON VÀ TXT, THAY BẰNG EXCEL
    """
    try:
        # 1. Kiểm tra ảnh đã xử lý
        if processed_image is None or processed_image.size == 0:
            messagebox.showwarning("Cảnh báo", "Không có ảnh đã xử lý để lưu!")
            return None

        # 2. Kiểm tra kết quả detection
        if not detections or len(detections) == 0:
            messagebox.showwarning("Cảnh báo", "Không có kết quả phát hiện để lưu!")
            return None

        # 3. Tạo timestamp và base_name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"fruit_analysis_{timestamp}"

        # 4. Chọn thư mục lưu
        output_dir = filedialog.askdirectory(
            title="Chọn thư mục lưu kết quả",
            initialdir=str(OUTPUTS_DIR)
        )
        if not output_dir:
            return None
        os.makedirs(output_dir, exist_ok=True)

        results = {}

        # 5. Lưu ảnh đã xử lý
        image_path = os.path.join(output_dir, f"{base_name}_processed.jpg")
        if not cv2.imwrite(image_path, processed_image):
            raise ValueError(f"Không thể lưu ảnh: {image_path}")
        results['image_path'] = image_path
        print(f"✅ Đã lưu ảnh: {image_path}")

        # 6. Lưu ảnh gốc nếu có
        if original_image_path and os.path.exists(original_image_path):
            import shutil
            original_copy_path = os.path.join(output_dir, f"{base_name}_original.jpg")
            shutil.copy2(original_image_path, original_copy_path)
            results['original_image_path'] = original_copy_path
            print(f"✅ Đã lưu ảnh gốc: {original_copy_path}")

        # 7. Tạo và lưu Excel thay cho JSON và TXT
        excel_path = os.path.join(output_dir, f"{base_name}_results.xlsx")
        if export_to_excel(detections, settings, excel_path):
            results['excel_path'] = excel_path
            print(f"✅ Đã lưu Excel: {excel_path}")
        else:
            print("⚠️ Không thể lưu file Excel")

        # 8. Tạo overlay
        try:
            overlay_path = os.path.join(output_dir, f"{base_name}_overlay.jpg")
            if create_overlay_image(processed_image, detections, overlay_path):
                results['overlay_path'] = overlay_path
                print(f"✅ Đã lưu overlay: {overlay_path}")
        except Exception as e:
            print(f"⚠️ Không tạo được overlay: {e}")

        # 9. Thông báo thành công
        file_count = len([k for k in results.keys() if k.endswith('_path')])
        messagebox.showinfo(
            "Thành công",
            f"✅ Đã lưu {file_count} file kết quả vào:\n{output_dir}\n\n"
            f"• {os.path.basename(image_path)}\n"
            f"• {os.path.basename(excel_path)}"
        )

        return results

    except Exception as e:
        import traceback
        traceback.print_exc()
        messagebox.showerror("Lỗi", f"Không thể lưu file: {e}")
        return None


def export_to_excel(detections, settings, excel_path):
    """Xuất dữ liệu sang Excel - Phiên bản mở rộng thay thế JSON và TXT"""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # TÍNH TOÁN THỐNG KÊ
        total_count = len(detections)

        # Đếm theo LOẠI SẢN PHẨM
        class_counts = {}
        class_quality_counts = {}
        class_scores = {}  # Lưu điểm của từng loại

        # Đếm theo chất lượng
        quality_counts = {}
        quality_scores = {}

        for det in detections:
            # Loại sản phẩm
            class_name = det.get('class', 'unknown')
            class_counts[class_name] = class_counts.get(class_name, 0) + 1

            # Lưu điểm cho từng loại
            if class_name not in class_scores:
                class_scores[class_name] = []
            class_scores[class_name].append(det.get('quality_score', 0))

            # Chất lượng
            quality = det.get('quality', 'unknown')
            quality_counts[quality] = quality_counts.get(quality, 0) + 1

            # Điểm chất lượng
            if quality not in quality_scores:
                quality_scores[quality] = []
            quality_scores[quality].append(det.get('quality_score', 0))

            # Thống kê chất lượng theo từng loại
            if class_name not in class_quality_counts:
                class_quality_counts[class_name] = {}
            if quality not in class_quality_counts[class_name]:
                class_quality_counts[class_name][quality] = 0
            class_quality_counts[class_name][quality] += 1

        # Tính điểm trung bình
        avg_scores = {}
        for quality, scores in quality_scores.items():
            avg_scores[quality] = sum(scores) / len(scores) if scores else 0

        # Tính điểm trung bình cho từng loại
        class_avg_scores = {}
        for class_name, scores in class_scores.items():
            class_avg_scores[class_name] = sum(scores) / len(scores) if scores else 0

        # Đếm sản phẩm chất lượng (chín)
        quality_good = quality_counts.get('ripe', 0) + quality_counts.get('good', 0)
        defect_count = quality_counts.get('bad', 0) + quality_counts.get('rotten', 0)
        defect_rate = (defect_count / total_count * 100) if total_count > 0 else 0

        # Tổng điểm trung bình
        total_scores = [det.get('quality_score', 0) for det in detections]
        avg_total_score = sum(total_scores) / len(total_scores) if total_scores else 0

        # TẠO DATAFRAME CHO CHI TIẾT (BỎ CỘT KÍCH THƯỚC)
        detection_data = []
        for i, det in enumerate(detections, 1):
            detection_data.append({
                'STT': i,
                'Loại sản phẩm': PRODUCT_NAMES_VI.get(det['class'], det['class']),
                'Class (English)': det['class'],
                'Chất lượng': QUALITY_NAMES_VI.get(det['quality'], det['quality']),
                'Quality (English)': det['quality'],
                'Điểm chất lượng': det['quality_score']
            })

        df_detections = pd.DataFrame(detection_data)

        # TẠO DATAFRAME CHO THỐNG KÊ SỐ LƯỢNG THEO LOẠI (CHI TIẾT)
        class_stats_data = []
        for class_name, count in class_counts.items():
            class_name_vn = PRODUCT_NAMES_VI.get(class_name, class_name)
            percentage = (count / total_count * 100) if total_count > 0 else 0
            avg_score = class_avg_scores.get(class_name, 0)

            # Thống kê chất lượng cho từng loại
            quality_details = []
            if class_name in class_quality_counts:
                for quality, q_count in class_quality_counts[class_name].items():
                    quality_vn = QUALITY_NAMES_VI.get(quality, quality)
                    q_percentage = (q_count / count * 100) if count > 0 else 0
                    quality_details.append(f"{quality_vn}: {q_count} ({q_percentage:.1f}%)")

            class_stats_data.append({
                'Loại sản phẩm': class_name_vn,
                'Class (English)': class_name,
                'Số lượng': count,
                'Tỷ lệ (%)': f"{percentage:.1f}%",
                'Điểm chất lượng TB': f"{avg_score:.2f}",
                'Phân bố chất lượng': ', '.join(quality_details)
            })

        df_class_stats = pd.DataFrame(class_stats_data)

        # TẠO DATAFRAME TỔNG HỢP SỐ LƯỢNG LOẠI (DẠNG BẢNG ĐƠN GIẢN)
        class_summary_data = []
        for class_name, count in sorted(class_counts.items(), key=lambda x: x[1], reverse=True):
            class_name_vn = PRODUCT_NAMES_VI.get(class_name, class_name)
            percentage = (count / total_count * 100) if total_count > 0 else 0
            avg_score = class_avg_scores.get(class_name, 0)

            class_summary_data.append({
                'STT': len(class_summary_data) + 1,
                'Loại sản phẩm': class_name_vn,
                'Số lượng': count,
                'Tỷ lệ (%)': f"{percentage:.1f}%",
                'Điểm TB': f"{avg_score:.2f}"
            })

        df_class_summary = pd.DataFrame(class_summary_data)

        # TẠO DATAFRAME CHO THỐNG KÊ CHẤT LƯỢNG
        stats_quality = []
        for quality, count in quality_counts.items():
            quality_vn = QUALITY_NAMES_VI.get(quality, quality)
            percentage = (count / total_count * 100) if total_count > 0 else 0
            avg_score = avg_scores.get(quality, 0)
            stats_quality.append({
                'Chỉ số': 'Chất lượng',
                'Loại': quality_vn,
                'Số lượng': count,
                'Tỷ lệ (%)': f"{percentage:.1f}%",
                'Điểm TB': f"{avg_score:.2f}"
            })

        df_stats_quality = pd.DataFrame(stats_quality)

        # TẠO DATAFRAME CHO TỔNG HỢP
        summary_data = [{
            'Thông số': 'Tổng số sản phẩm',
            'Giá trị': total_count,
            'Đơn vị': 'sản phẩm'
        }, {
            'Thông số': 'Số loại sản phẩm',
            'Giá trị': len(class_counts),
            'Đơn vị': 'loại'
        }]

        # Thêm số lượng từng loại vào tổng hợp
        for class_name, count in sorted(class_counts.items(), key=lambda x: x[1], reverse=True):
            class_name_vn = PRODUCT_NAMES_VI.get(class_name, class_name)
            percentage = (count / total_count * 100) if total_count > 0 else 0
            summary_data.append({
                'Thông số': f'Số lượng {class_name_vn}',
                'Giá trị': count,
                'Đơn vị': 'sản phẩm',
                'Tỷ lệ': f"{percentage:.1f}%"
            })

        summary_data.extend([
            {
                'Thông số': 'Sản phẩm chất lượng (chín/tốt)',
                'Giá trị': quality_good,
                'Đơn vị': 'sản phẩm'
            }, {
                'Thông số': 'Tỷ lệ hỏng',
                'Giá trị': f"{defect_rate:.1f}%",
                'Đơn vị': ''
            }, {
                'Thông số': 'Điểm chất lượng trung bình',
                'Giá trị': f"{avg_total_score:.2f}",
                'Đơn vị': '/1.0'
            }, {
                'Thông số': 'Loại sản phẩm chính',
                'Giá trị': PRODUCT_NAMES_VI.get(settings.get('product_type', 'auto'),
                                                settings.get('product_type', 'auto')),
                'Đơn vị': ''
            }, {
                'Thông số': 'Ngưỡng tin cậy',
                'Giá trị': f"{settings.get('confidence_threshold', 0.5):.2f}",
                'Đơn vị': ''
            }, {
                'Thông số': 'Thời gian phân tích',
                'Giá trị': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Đơn vị': ''
            }
        ])

        df_summary = pd.DataFrame(summary_data)

        # GHI RA EXCEL VỚI ĐỊNH DẠNG
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Chi tiết phân loại
            df_detections.to_excel(writer, sheet_name='Chi tiết phân loại', index=False)

            # Sheet 2: Tổng hợp số lượng theo loại
            df_class_summary.to_excel(writer, sheet_name='Số lượng từng loại', index=False)

            # Sheet 3: Thống kê chi tiết theo loại
            df_class_stats.to_excel(writer, sheet_name='Thống kê chi tiết từng loại', index=False)

            # Sheet 4: Thống kê chất lượng
            df_stats_quality.to_excel(writer, sheet_name='Thống kê chất lượng', index=False)

            # Sheet 5: Tổng hợp
            df_summary.to_excel(writer, sheet_name='Tổng hợp', index=False)

            # Lấy workbook để định dạng
            workbook = writer.book

            # Định dạng các sheet
            sheets_titles = {
                'Chi tiết phân loại': "CHI TIẾT PHÂN LOẠI TỪNG SẢN PHẨM",
                'Số lượng từng loại': "SỐ LƯỢNG TỪNG LOẠI SẢN PHẨM",
                'Thống kê chi tiết từng loại': "THỐNG KÊ CHI TIẾT TỪNG LOẠI SẢN PHẨM",
                'Thống kê chất lượng': "THỐNG KÊ PHÂN LOẠI THEO CHẤT LƯỢNG",
                'Tổng hợp': "BÁO CÁO TỔNG HỢP PHÂN TÍCH"
            }

            for sheet_name, title in sheets_titles.items():
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    _format_excel_sheet(ws, title)

            # Thêm bảng xếp hạng loại sản phẩm vào sheet tổng hợp
            ws_total = workbook['Tổng hợp']
            ws_total.append([])
            ws_total.append(["BẢNG XẾP HẠNG SỐ LƯỢNG THEO LOẠI:"])

            # Thêm header cho bảng xếp hạng
            ws_total.append(["Hạng", "Loại sản phẩm", "Số lượng", "Tỷ lệ", "Điểm chất lượng TB"])

            # Thêm dữ liệu xếp hạng
            rank = 1
            for class_name, count in sorted(class_counts.items(), key=lambda x: x[1], reverse=True):
                class_name_vn = PRODUCT_NAMES_VI.get(class_name, class_name)
                percentage = (count / total_count * 100) if total_count > 0 else 0
                avg_score = class_avg_scores.get(class_name, 0)
                ws_total.append([rank, class_name_vn, count, f"{percentage:.1f}%", f"{avg_score:.2f}"])
                rank += 1

            # Thêm phân tích chi tiết từng loại
            ws_total.append([])
            ws_total.append(["PHÂN TÍCH CHI TIẾT TỪNG LOẠI:"])

            for class_name, count in sorted(class_counts.items(), key=lambda x: x[1], reverse=True):
                class_name_vn = PRODUCT_NAMES_VI.get(class_name, class_name)
                percentage = (count / total_count * 100) if total_count > 0 else 0
                avg_score = class_avg_scores.get(class_name, 0)

                # Phân tích chất lượng
                if class_name in class_quality_counts:
                    good_count = class_quality_counts[class_name].get('ripe', 0) + class_quality_counts[class_name].get(
                        'good', 0)
                    bad_count = class_quality_counts[class_name].get('bad', 0) + class_quality_counts[class_name].get(
                        'rotten', 0)
                    good_rate = (good_count / count * 100) if count > 0 else 0
                    bad_rate = (bad_count / count * 100) if count > 0 else 0

                    ws_total.append([f"• {class_name_vn} ({count} sản phẩm, {percentage:.1f}%)"])
                    ws_total.append([f"  - Điểm chất lượng TB: {avg_score:.2f}/1.0"])
                    ws_total.append([f"  - Chất lượng tốt: {good_count} ({good_rate:.1f}%)"])
                    ws_total.append([f"  - Chất lượng hỏng: {bad_count} ({bad_rate:.1f}%)"])
                    ws_total.append([])

            ws_total.append(["KẾT LUẬN TỔNG QUAN:"])

            if defect_rate < 5:
                conclusion = "✅ CHẤT LƯỢNG TỐT - Tỷ lệ hỏng thấp (<5%). Sản phẩm đạt yêu cầu xuất khẩu."
            elif defect_rate < 20:
                conclusion = "⚠️ CHẤT LƯỢNG TRUNG BÌNH - Tỷ lệ hỏng vừa phải (5-20%). Cần kiểm tra và phân loại lại."
            else:
                conclusion = "❌ CHẤT LƯỢNG KÉM - Tỷ lệ hỏng cao (>20%). Cần xử lý và loại bỏ sản phẩm hỏng."

            ws_total.append([conclusion])
            ws_total.append([])
            ws_total.append([f"Thời gian tạo báo cáo: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws_total.append(["Hệ thống phân loại sản phẩm nông nghiệp - Phiên bản 1.0"])

        print(f"✅ File Excel đã được tạo với {len(detections)} bản ghi, {len(class_counts)} loại sản phẩm")
        return True

    except ImportError:
        messagebox.showerror("Lỗi",
                             "Thiếu thư viện pandas hoặc openpyxl. Vui lòng cài đặt:\npip install pandas openpyxl")
        return False
    except Exception as e:
        print(f"❌ Lỗi export Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

def _format_excel_sheet(worksheet, title):
    """Định dạng worksheet Excel"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Thiết lập độ rộng cột tự động
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Định dạng tiêu đề
    if worksheet.max_row > 0:
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Thêm border cho tất cả cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                continue
            if cell.column == 1:  # Cột STT
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.alignment = Alignment(horizontal="right")

    # Thêm title
    if title:
        worksheet.insert_rows(1)
        worksheet.merge_cells(f'A1:{get_column_letter(worksheet.max_column)}1')
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")


def create_overlay_image(image, detections, output_path):
    """Tạo ảnh overlay với thông tin chi tiết"""
    try:
        overlay = image.copy()
        h, w = image.shape[:2]

        # Đặt màu cho các loại chất lượng
        quality_colors = {
            'ripe': (0, 255, 0),      # Xanh lá - Chín
            'good': (0, 200, 100),    # Xanh lá nhạt - Tốt
            'unripe': (255, 255, 0),  # Vàng - Xanh
            'bad': (0, 165, 255),     # Cam - Hỏng nhẹ
            'rotten': (0, 0, 255)     # Đỏ - Hỏng nặng
        }

        # Vẽ bounding boxes và thông tin
        for i, det in enumerate(detections, 1):
            bbox = det['bbox']
            quality = det.get('quality', 'unknown')
            color = quality_colors.get(quality, (255, 255, 255))

            # Vẽ bounding box
            cv2.rectangle(overlay,
                         (int(bbox[0]), int(bbox[1])),
                         (int(bbox[2]), int(bbox[3])),
                         color, 2)

            # Tạo label
            label = f"{i}. {PRODUCT_NAMES_VI.get(det['class'], det['class'])} - {QUALITY_NAMES_VI.get(quality, quality)}"

            # Vẽ background cho label
            label_size, baseline = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)
            cv2.rectangle(overlay,
                         (int(bbox[0]), int(bbox[1]) - label_size[1] - 5),
                         (int(bbox[0]) + label_size[0], int(bbox[1])),
                         color, -1)

            # Vẽ text
            cv2.putText(overlay, label,
                       (int(bbox[0]), int(bbox[1]) - 5),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

        # Thêm watermark và thông tin
        cv2.putText(overlay, f"Tổng: {len(detections)} sản phẩm",
                   (10, 30), cv2.FONT_HERSHEY_DUPLEX, 0.8, (0, 255, 0), 2)

        # Thêm legend chất lượng
        y_offset = 60
        for quality, color in quality_colors.items():
            if quality in QUALITY_NAMES_VI:
                label = f"{QUALITY_NAMES_VI[quality]}"
                cv2.rectangle(overlay, (10, y_offset-15), (25, y_offset), color, -1)
                cv2.putText(overlay, label,
                           (30, y_offset),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
                y_offset += 25

        # Thêm timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        cv2.putText(overlay, timestamp,
                   (w - 200, h - 10),
                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

        # Lưu ảnh
        cv2.imwrite(output_path, overlay)
        return True
    except Exception as e:
        print(f"⚠️ Không tạo được overlay: {e}")
        return False


def get_file_size(file_path):
    """Lấy kích thước file"""
    try:
        size_bytes = os.path.getsize(file_path)
        # Chuyển đổi sang đơn vị phù hợp
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.2f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.2f} TB"
    except:
        return "Unknown"


def list_output_files(output_dir=None):
    """Liệt kê các file kết quả trong thư mục"""
    if output_dir is None:
        output_dir = OUTPUTS_DIR

    results = []
    try:
        for root, dirs, files in os.walk(output_dir):
            for file in files:
                if file.endswith(('.jpg', '.png', '.xlsx')):
                    file_path = os.path.join(root, file)
                    file_info = {
                        'name': file,
                        'path': file_path,
                        'size': get_file_size(file_path),
                        'modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M')
                    }
                    results.append(file_info)

        # Sắp xếp theo thời gian sửa đổi (mới nhất trước)
        results.sort(key=lambda x: x['modified'], reverse=True)
        return results

    except Exception as e:
        print(f"Lỗi liệt kê file: {e}")
        return []


def load_results(json_path):
    """Tải kết quả từ file JSON (cho tương thích)"""
    try:
        import json
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data
    except Exception as e:
        print(f"⚠️ Lỗi tải file JSON: {e}")
        # Trả về dict rỗng để tương thích
        return {
            'detections': [],
            'statistics': {},
            'settings': {},
            'timestamp': ''
        }